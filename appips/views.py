from django.shortcuts import render
from django.views.generic import CreateView, DetailView
from django.urls import reverse_lazy
from django.contrib import messages
from .models import *
from datetime import date
from django.shortcuts import get_object_or_404 , redirect
from django.http import HttpResponse,JsonResponse
from .forms import *
import openpyxl
from openpyxl.styles import Font

#Home
def home(request):
    return render(request, 'menu.html')

#Clase para crear usuario

class UsuarioCreate(CreateView):
    model = Usuario
    fields = '__all__'  
    template_name = 'usuario_form.html'  
    success_url = reverse_lazy('listar_usuarios') 

    def form_valid(self, form):
        messages.success(self.request, "El usuario ha sido creado y guardado exitosamente.")
        return super().form_valid(form)
    
#Función para buscar usuarios

def buscar_usuario(request):
    criterio = request.GET.get('criterio', '')  # Obtén el criterio de búsqueda o un string vacío
    if criterio:
        # Realiza la búsqueda basada en el criterio
        usuarios = Usuario.objects.filter(
            numero_identificacion__icontains=criterio
        ) | Usuario.objects.filter(
            primer_nombre__icontains=criterio
        ) | Usuario.objects.filter(
            primer_apellido__icontains=criterio
        )
    else:
        # Lista todos los usuarios si no hay criterio
        usuarios = Usuario.objects.all()

    return render(request, 'listar_usuarios.html', {
        'resultados': usuarios,
        'criterio': criterio,
        'request': request
    })
    
def crear_registro(request, numero_identificacion):
    # Obtener el usuario correspondiente
    usuario = get_object_or_404(Usuario, numero_identificacion=numero_identificacion)
    
    # Verificar si ya existe un registro para el usuario
    if Registro.objects.filter(id_usuario=usuario).exists():
        return JsonResponse({"mensaje": "Error: Ya existe un registro asociado a este usuario."}, status=400)
    # Crear el registro asociado
    registro = Registro.objects.create(
            id_usuario=usuario,
            tipo_registro=2,
            codigo_habilitacion_ips_primaria=528380048301)
   
    # Devolver respuesta en formato JSON
    return JsonResponse({"mensaje": "Registro creado exitosamente"})


#Editar registro
class DetalleRegistroView(DetailView):
    template_name = 'detalle_registro.html'
    model = Registro
    context_object_name = 'registro'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registro = self.object
        usuario = registro.id_usuario
        # Calcular la edad
        if usuario.fecha_nacimiento:
            today = date.today()
            edad = today.year - usuario.fecha_nacimiento.year - (
                (today.month, today.day) < (usuario.fecha_nacimiento.month, usuario.fecha_nacimiento.day)
            )
        else:
            edad = None

        # Datos básicos del usuario
        context['usuario'] = usuario
        context['edad'] = edad

        # Agregar Agudeza Visual
        agudeza_visual = AgudezaVisual.objects.filter(consecutivo_registro=registro).first()
        if agudeza_visual:
            context['agudeza_visual'] = agudeza_visual
        else:
            context['form_agudeza_visual'] = AgudezaVisualForm(usuario=usuario)

        # Agregar Anticoncepción
        anticoncepcion = Anticoncepcion.objects.filter(consecutivo_registro=registro).first()
        if anticoncepcion:
            context['anticoncepcion'] = anticoncepcion
        else:
            context['form_anticoncepcion'] = AnticoncepcionForm()

       # Agregar Cancer de cervix
        ca_cervix = Ca_Cervix.objects.filter(consecutivo_registro=registro).first()
        if ca_cervix:
            context['ca_cervix'] = ca_cervix
        else:
            context['form_ca_cervix'] = CaCervixForm()

        # Agregar Cancer de colon
        ca_colon = Ca_Colon.objects.filter(consecutivo_registro=registro).first()
        if ca_colon:
            context['ca_colon'] = ca_cervix
        else:
            context['form_ca_colon'] = CaProstataForm()

        # Agregar Cancer de próstata
        ca_prostata = Ca_Prostata.objects.filter(consecutivo_registro=registro).first()

        if ca_prostata:
            context['ca_prostata'] = ca_prostata
        else:
            context['form_ca_prostata'] = CaProstataForm()
       
        ca_mama=Ca_Mama.objects.filter(consecutivo_registro=registro).first()

        if ca_mama:
            context['ca_mama'] = ca_mama
        else:
            context['form_ca_mama'] = CaMamaForm()
        
        # Agregar Salud Bucal
        salud_bucal=Salud_Bucal.objects.filter(consecutivo_registro=registro).first()

        if salud_bucal:
            context['salud_bucal'] = salud_bucal
        else:
            context['form_salud_bucal'] = SaludBucalForm()
        
        # Agregar Gestación.
        gestacion=Gestacion.objects.filter(consecutivo_registro=registro).first()

        if gestacion:
            context['gestacion'] = gestacion
        else:
            context['form_gestacion'] = GestacionForm()

        #Agregar Gestación menores.

        gestacion_menores=Gestacion_menores_hasta_siente_meses.objects.filter(consecutivo_registro=registro).first()

        if gestacion_menores:
            context['gestacion_menores'] = gestacion_menores
        else:
            context['form_gestacion_menores'] = GestacionmenoresForm()

        no_reporte=No_Reporte.objects.filter(consecutivo_registro=registro).first()

        if no_reporte:
            context['no_reporte'] = no_reporte
        else:
            context['form_no_reporte'] = NoReporteForm()

        primera_infacia=Primera_Infancia.objects.filter(consecutivo_registro=registro).first()

        if  primera_infacia:
            context['primera_infacia'] = primera_infacia
        else:
            context['form_primera_infacia'] =  PrimeraInfanciaForm()

        recien_nacido=Recien_Nacidos.objects.filter(consecutivo_registro=registro).first()

        if  recien_nacido:
            context['recien_nacido'] = recien_nacido
        else:
            context['form_recien_nacido'] = RecienNacidoForm()

        riesgo_cardiovascular=Riesgo_Cardiovascular.objects.filter(consecutivo_registro=registro).first()

        if  riesgo_cardiovascular:
            context['riesgo_cardiovascular'] = riesgo_cardiovascular
        else:
            context['form_riesgo_cardiovascular'] = RiesgoCardiovascularForm()


        test_o_12=Test_0_12.objects.filter(consecutivo_registro=registro).first()

        if  test_o_12:
            context['test_o_12'] = test_o_12
        else:
            context['form_test_o_12'] = Test_0_12Form()

        test_o_7=Test_0_7.objects.filter(consecutivo_registro=registro).first()

        if  test_o_7:
            context['test_o_7'] = test_o_7
        else:
            context['form_test_o_7'] = Test_0_7Form()

        test_mayores_60=Test_Mayores_60.objects.filter(consecutivo_registro=registro).first()

        if  test_mayores_60:
            context['test_mayores_60'] = test_mayores_60
        else:
            context['form_test_mayores_60'] = Test_Mayores_60Form()

        toda_poblacion=Toda_poblacion.objects.filter(consecutivo_registro=registro).first()

        if  toda_poblacion:
            context['toda_poblacion'] = toda_poblacion
        else:
            context['form_toda_poblacion'] = Toda_poblacionForm()

        tuberculosis=Tuberculosis.objects.filter(consecutivo_registro=registro).first()

        if  tuberculosis:
            context['tuberculosis'] = tuberculosis
        else:
            context['form_tuberculosis'] = TuberculosisForm()


        return context

    def post(self, request, *args, **kwargs):
        registro = self.get_object()

        # Formularios
        form_agudeza_visual = AgudezaVisualForm(request.POST)
        form_anticoncepcion = AnticoncepcionForm(request.POST)
        form_ca_cervix = CaCervixForm(request.POST)
        form_ca_colon = CaColonForm(request.POST)
        form_ca_prostata=CaProstataForm(request.POST)
        form_ca_mama=CaMamaForm(request.POST)
        form_salud_bucal=SaludBucalForm(request.POST)
        form_gestacion=GestacionForm(request.POST)
        form_gestacion_menores=GestacionmenoresForm(request.POST)
        form_no_reporte=NoReporteForm(request.POST)
        form_primera_infacia=PrimeraInfanciaForm(request.POST)
        form_recien_nacido=RecienNacidoForm(request.POST)
        form_riesgo_cardiovascular=RiesgoCardiovascularForm(request.POST)
        form_test_o_12 =Test_0_12Form(request.POST)
        form_test_o_7 =Test_0_7Form(request.POST)
        form_test_mayores_60= Test_0_7Form(request.POST)
        form_toda_poblacion=Toda_poblacionForm(request.POST)
        form_tuberculosis=TuberculosisForm(request.POST)

        # Guardar Agudeza Visual
        if form_agudeza_visual.is_valid():
            agudeza_visual = form_agudeza_visual.save(commit=False)
            agudeza_visual.consecutivo_registro = registro
            agudeza_visual.save()
            return redirect('detalle_registro', pk=registro.pk)

        # Guardar Anticoncepción
        if form_anticoncepcion.is_valid():
            anticoncepcion = form_anticoncepcion.save(commit=False)
            anticoncepcion.consecutivo_registro = registro
            anticoncepcion.save()
            return redirect('detalle_registro', pk=registro.pk)
        
        if form_ca_cervix.is_valid():
            ca_cervix = form_ca_cervix.save(commit=False)
            ca_cervix.consecutivo_registro = registro
            ca_cervix.save()
            return redirect('detalle_registro', pk=registro.pk)
        
        # Guardar cancer de colon

        if form_ca_colon.is_valid():
            ca_colon = form_ca_colon.save(commit=False)
            ca_colon.consecutivo_registro = registro
            ca_colon.save()
            return redirect('detalle_registro', pk=registro.pk)
        
        # Guardar cancer de prostata

        if form_ca_prostata.is_valid():
            ca_prostata = form_ca_prostata.save(commit=False)
            ca_prostata.consecutivo_registro = registro
            ca_prostata.save()
            return redirect('detalle_registro', pk=registro.pk)
        
        # Guardar cancer de mama
        if form_ca_mama.is_valid():
            ca_mama = form_ca_mama.save(commit=False)
            ca_mama.consecutivo_registro = registro
            ca_mama.save()
            return redirect('detalle_registro', pk=registro.pk)
        
        # Guardar getación
        if form_salud_bucal.is_valid():
            salud_bucal=form_salud_bucal.save(commit=False)
            salud_bucal.consecutivo_registro = registro
            salud_bucal.save()
            return redirect('detalle_registro', pk=registro.pk)
        
        if form_gestacion.is_valid():
            gestacion=form_gestacion.save(commit=False)
            gestacion.consecutivo_registro = registro
            gestacion.save()
            return redirect('detalle_registro', pk=registro.pk)
        
        if form_gestacion_menores.is_valid():
            gestacion_menores=form_gestacion_menores.save(commit=False)
            gestacion_menores.consecutivo_registro = registro
            gestacion_menores.save()
            return redirect('detalle_registro', pk=registro.pk)
        
        if form_no_reporte.is_valid():
            no_reporte=form_no_reporte.save(commit=False)
            no_reporte.consecutivo_registro = registro
            no_reporte.save()
            return redirect('detalle_registro', pk=registro.pk)
        
        if  form_primera_infacia.is_valid():
            primera_infacia= form_primera_infacia.save(commit=False)
            primera_infacia.consecutivo_registro = registro
            primera_infacia.save()
            return redirect('detalle_registro', pk=registro.pk)
        

        if  form_recien_nacido.is_valid():
            recien_nacido= form_recien_nacido.save(commit=False)
            recien_nacido.consecutivo_registro = registro
            recien_nacido.save()
            return redirect('detalle_registro', pk=registro.pk)
        

        if form_riesgo_cardiovascular.is_valid():
           riesgo_cardiovascular =form_riesgo_cardiovascular.save(commit=False)
           riesgo_cardiovascular.consecutivo_registro = registro
           riesgo_cardiovascular.save()
           return redirect('detalle_registro', pk=registro.pk)
        

        if form_test_o_12.is_valid():
           test_o_12 =form_test_o_12.save(commit=False)
           test_o_12.consecutivo_registro = registro
           test_o_12.save()
           return redirect('detalle_registro', pk=registro.pk)
        
        if form_test_o_7.is_valid():
           test_o_7 =form_test_o_7.save(commit=False)
           test_o_7.consecutivo_registro = registro
           test_o_7.save()
           return redirect('detalle_registro', pk=registro.pk)
        

        #form_test_mayores_60
        if form_test_mayores_60.is_valid():
           test_mayores_60 =form_test_mayores_60.save(commit=False)
           test_mayores_60.consecutivo_registro = registro
           test_mayores_60.save()
           return redirect('detalle_registro', pk=registro.pk)
        
        if form_toda_poblacion.is_valid():
           toda_poblacion=form_toda_poblacion.save(commit=False)
           toda_poblacion.consecutivo_registro = registro
           toda_poblacion.save()
           return redirect('detalle_registro', pk=registro.pk)
        
        #form_tuberculosis
        
        if form_tuberculosis.is_valid():
           tuberculosis=form_tuberculosis.save(commit=False)
           tuberculosis.consecutivo_registro = registro
           tuberculosis.save()
           return redirect('detalle_registro', pk=registro.pk)

                 
        # Si algún formulario falla, devolver los errores al template
        return self.render_to_response({
            'registro': registro,
            'form_agudeza_visual': form_agudeza_visual,
            'form_anticoncepcion': form_anticoncepcion,
            'form_ca_cervix': form_ca_cervix,
            'form_ca_colon': form_ca_colon,
            'form_ca_prostata':form_ca_prostata,
            'form_ca_mama':form_ca_mama,
            'form_salud_bucal':form_salud_bucal,
            'form_gestacion':form_gestacion,
            'form_gestacion_menores':form_gestacion_menores,
            'form_no_reporte':form_no_reporte,
            'form_primera_infacia':form_primera_infacia,
            'form_recien_nacido':form_recien_nacido,
            'riesgo_cardiovascular':riesgo_cardiovascular,
            'form_test_o_12':form_test_o_12,
            'form_test_o_7':form_test_o_7,
            'form_test_o_7':form_test_o_7,
            'form_toda_poblacion':form_toda_poblacion,
            'form_tuberculosis':form_tuberculosis

        })
    

#     encabezados = [
#         "Tipo de registro", "Consecutivo de registro", "Código de habilitación IPS primaria",
#         "Tipo de identificación", "Número de identificación del usuario", "Primer apellido del usuario",
#         "Segundo apellido del usuario", "Primer nombre del usuario", "Segundo nombre del usuario",
#         "Fecha de Nacimiento", "Sexo", "Código pertenencia étnica",
#         "Código de ocupación (toda la población)", "Código de nivel educativo (toda la población)",
#         "Gestante", "Sífilis Gestacional o congénita", "Resultado de la prueba mini-mental state",
#         "Hipotiroidismo Congénito", "Sintomático respiratorio (toda la población)", 
#         "Sintomático respiratorio (toda la población)","Consumo de tabaco","Lepra",
#         "Obesidad o Desnutrición Proteico Calórica","Resultado del tacto rectal","Acido fólico preconcepcional",
#         ,"Enfermedad Mental",,
#         "Agudeza visual lejana ojo izquierdo ","Agudeza visual lejana ojo derecho ","Fecha del peso ","Peso en Kilogramos ",
#         "Fecha de la talla","Talla en centímetros ","Fecha probable de parto","Código país","Clasificación del riesgo gestacional",
#         "Resultado de colonoscopia tamizaje","Resultado de tamizaje auditivo neonatal","DPT menores de 5 años","Resultado de tamizaje VALE",
#         "Neumococo","Resultado de tamizaje para hepatitis C","Resultado de escala abreviada de desarrollo área de motricidad gruesa",
#         "Resultado de escala abreviada de desarrollo área de motricidad finoadaptativa","Resultado de escala abreviada de desarrollo área personal social",
#         "Resultado de escala abreviada de desarrollo área de motricidad audición lenguaje",
#         "Fecha de atención parto o cesárea","Fecha de salida de atención parto o cesárea",","Fecha de consulta de valoración integral",
#         ,"Fecha de primera consulta prenatal",,
#         "Fecha de último control prenatal de seguimiento","Suministro de ácido fólico en el control prenatal durante el periodo reportado","Suministro de sulfato ferroso en el control prenatal durante el periodo reportado",
#         "Suministro de carbonato de calcio en el control prenatal durante el periodo reportado","Fecha de valoración agudeza visual","Fecha de tamizaje VALE","Fecha del tacto rectal","Fecha de tamización con oximetría pre y post ductal",
#         "Fecha de realización colonoscopia tamizaje","Fecha de la prueba sangre oculta en materia fecal (tamizaje Ca de colon)","Consulta de Psicología","Fecha de tamizaje auditivo neonatal","Suministro de  fortificación casera en la primera infancia (6 a 23 meses)",
#         "Suministro de vitamina A en la primera infancia (24 a 60 meses)","Fecha de toma LDL","Fecha de toma PSA","Preservativos entregados a pacientes con ITS","Fecha de tamizaje visual neonatal","Fecha de atención en salud bucal por profesional en odontología",
#         "Suministro de hierro en la primera Infancia (24 a 59 meses)","Fecha de antígeno de superficie hepatitis B (Toda la población)","Resultado de antígeno de superficie hepatitis B (Toda la población)","Fecha de toma de prueba tamizaje para sífilis",
#         "Resultado de prueba tamizaje para sífilis","Fecha de toma de prueba para VIH","Resultado de prueba para VIH","Fecha de TSH neonatal","Resultado de TSH neonatal","Tamizaje del cáncer de cuello uterino ","Fecha de tamizaje cáncer de cuello uterino",
#         "Resultado tamizaje cáncer de cuello uterino","Calidad en la muestra de citología cervicouterina","Código de habilitación IPS donde se realiza  tamizaje cáncer de cuello uterino","Fecha de colposcopia","Resultado de LDL","Fecha de biopsia cervicouterina",
#         "Resultado de biopsia cervicouterina","Resultado de HDL","Fecha de toma de mamografía","Resultado de mamografía","Resultado de triglicéridos","Fecha de toma biopsia de mama","Fecha de resultado biopsia de mama","Resultado de biopsia de mama","COP por persona",
#         "Resultado de hemoglobina","Fecha de toma glicemia basal","Fecha de toma creatinina","Resultado de creatinina","Preservativos entregados a pacientes con ITS","Resultado de PSA","Fecha de toma  de tamizaje hepatitis C","Fecha de toma HDL","Fecha de toma de baciloscopia diagnóstico",
#         "Resultado de baciloscopia diagnóstico","Clasificación del riesgo cardiovascular", "Tratamiento para sífilis gestacional","Tratamiento para Sífilis Congénita","Clasificación del riesgo metabólico","Fecha de toma triglicéridos"
#     ]

def generar_reporte_excel(request):
    # Crear un archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Usuarios"

# Encabezados del reporte
    encabezados = [

#Datos registro

        "Tipo de registro", "Consecutivo de registro", "Código de habilitación IPS primaria",

#Datos usuarios

        "Tipo de Identificación","Número de Identificación  del usuario", "Primer Apellido del usuario",
        "Segundo Apellido del usuario", "Primer Nombre del usuario", "Segundo Nombre del usuario", "Fecha de Nacimiento",
        "Sexo", "Código Pertenencia Étnica", "Código Ocupación", "Código País","Código Nivel Educativo",
# Campos de Gestacion
    
        "Gestante", "Ácido Fólico Preconcepcional", "Fecha Probable de Parto",
        "Clasificación de Riesgo Gestacional", "Fecha de atención parto o cesárea ", 
        " Fecha de salida de atención parto o cesárea", "Fecha Primera Consulta Prenatal",
        "Fecha de último control prenatal de seguimiento", "Suministro Ácido Fólico", "Suministro de sulfato ferroso en el control prenatal durante el periodo reportado","Suministro de carbonato de calcio en el control prenatal durante el periodo reportado",
        
# Campos de Agudeza Visual
        "Agudeza Visual lejana Izquierdo", "Agudeza Visual lejana Derecho", "Fecha Valoración de Agudeza Visual",
 #Campos de Anticoncepcion 
        "Fecha de atención en salud para la asesoría en anticoncepción","Suministro de método anticonceptivo","Fecha de suministro de método anticonceptivo",
 # Campos de Ca_Cervix
        "Tamizaje del cáncer de cuello uterino","Tratamiento ablativo o de escisión posterior a la realización de la técnica de inspección visual (Crioterapia o LETZ)",
        "Fecha de tamizaje cáncer de cuello uterino","Resultado tamizaje cáncer de cuello uterino","Calidad en la muestra de citología cervicouterina","Código de habilitación IPS donde se realiza  tamizaje cáncer de cuello uterino","Fecha de colposcopia",
        "Fecha de biopsia cervicouterina","Resultado de biopsia cervicouterina" ,

 #Campos Gestacion menenores hasta 7 meses

        "Fecha de atención en salud para la promoción y apoyo de la lactancia materna",

   #Campos No reporte
  
        "Sífilis Gestacional o congénita","Lepra","Obesidad o Desnutrición Proteico Calórica","Enfermedad Mental",
        "Cáncer de Cérvix","DPT menores de 5 años","Neumococo","Consulta de Psicología","Preservativos entregados a pacientes con ITS","Fecha de hemoglobina","Tratamiento de sifilis gestacional",
        "Tratamiento de sifilis congenita",

  # Campos de Primera_Infancia
        "Suministro de  fortificación casera en la primera infancia (6 a 23 meses)",
        "Suministro de vitamina A en la primera infancia (24 a 60 meses)","Suministro de hierro en la primera Infancia (24 a 59 meses)",

   # Campos Recien_Nacidos
       "Resultado de tamizaje auditivo neonatal","Resultado de tamizaje visual  neonatal",
        "Resultado de tamización con oximetría pre y post ductal","Fecha de tamización con oximetría pre y post ductal","Fecha de tamizaje auditivo neonatal","Fecha de tamizaje visual neonatal",
         "Fecha de TSH neonatal","Resultado de TSH neonatal",

    # Datos Riesgo_Cardiovascular
        "Consumo de tabaco",
        "Resultado de glicemia basal",
        "Fecha de toma LDL",
        "Resultado de LDL",
        "Resultado de HDL",
        "Resultado de triglicéridos",
        "Fecha de toma de hemoglobina",
        "Resultado de hemoglobina", 
        "Fecha de toma glicemia basal",
        "Fecha de toma creatinina",
        "Resultado de creatinina",
        "Fecha de toma HDL",
        "Clasificación del riesgo cardiovascular",
        "Clasificación del riesgo metabólico",
        "Fecha de toma triglicéridos",

    #Campos tamizaje  0 a 12 años
        "Resultado de tamizaje VALE","Fecha de tamizaje VALE",

    #Campos tamizaje  0 a 7 años
        "Resultado de escala abreviada de desarrollo área de motricidad gruesa",
        "Resultado de escala abreviada de desarrollo área de motricidad finoadaptativa","Resultado de escala abreviada de desarrollo área personal social",
        "Resultado de escala abreviada de desarrollo área de motricidad audición lenguaje",

    #Campos Test Vejez mayores de 60 años           
        "Resultado de la prueba mini-mental state",

    # Campos Toda_poblacion
        "Fecha del peso ","Peso en Kilogramos ","Fecha de la talla","Talla en centímetros ","Resultado de tamizaje para hepatitis C","Fecha de consulta de valoración integral",
        "Fecha de antígeno de superficie hepatitis B (Toda la población)", "Resultado de antígeno de superficie hepatitis B (Toda la población)",
        "Fecha de toma de prueba tamizaje para sífilis","Resultado de prueba tamizaje para sífilis","Fecha de toma de prueba para VIH","Resultado de prueba para VIH",
        "Fecha de toma  de tamizaje hepatitis C",
        
    #Campos tuberculosis
        "Sintomático respiratorio (toda la población)",
        "Fecha de toma de baciloscopia diagnóstico",
        "Resultado de baciloscopia diagnóstico",

    #campos cancer de colon

        "Resultado de la prueba sangre oculta en materia fecal (tamizaje ", 
        "Resultado de colonoscopia tamizaje",
        "Fecha de realización colonoscopia tamizaje",
        "Fecha de la prueba sangre oculta en materia fecal (tamizaje Ca de colon)",

        #Campos cancer de prostata
        "Resultado del tacto rectal", "Fecha del tacto rectal","Fecha de toma PSA","Resultado de PSA",

        #Salud bucal

        "Fecha de atención en salud bucal por profesional en odontología","COP por persona",

        #Cancer de mama

        "Fecha de toma de mamografía","Resultado de mamografía","Fecha de toma biopsia de mama", "Fecha de resultado biopsia de mama","Resultado de biopsia de mama"
    ]

    # Escribir encabezados en la primera fila
    for col_num, header in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_num, value=header)
        celda.font = Font(bold=True)

    # Obtener datos de Usuario, Gestacion y AgudezaVisual
    usuarios = Usuario.objects.filter(registro__isnull=False).distinct()

    for row_num, usuario in enumerate(usuarios, 2):
        # Datos de Registro
        registro = Registro.objects.filter(id_usuario=usuario).first()
        if registro:
            ws.cell(row=row_num, column=1, value=registro.tipo_registro)
            ws.cell(row=row_num, column=2, value=registro.consecutivo_registro)
            ws.cell(row=row_num, column=3, value=registro.codigo_habilitacion_ips_primaria)
        else:
            ws.cell(row=row_num, column=1, value="N/A")
            ws.cell(row=row_num, column=2, value="N/A")
            ws.cell(row=row_num, column=3, value="N/A")

    # Escribir datos básicos de Usuario

        ws.cell(row=row_num, column=4, value=usuario.tipo_identificacion)
        ws.cell(row=row_num, column=5, value=usuario.numero_identificacion)
        ws.cell(row=row_num, column=6, value=usuario.primer_apellido)
        ws.cell(row=row_num, column=7, value=usuario.segundo_apellido)
        ws.cell(row=row_num, column=8, value=usuario.primer_nombre)
        ws.cell(row=row_num, column=9, value=usuario.segundo_nombre)
        ws.cell(row=row_num, column=10, value=usuario.fecha_nacimiento.strftime('%Y-%m-%d') if usuario.fecha_nacimiento else "N/A")
        ws.cell(row=row_num, column=11, value=usuario.sexo)
        ws.cell(row=row_num, column=12, value=usuario.codigo_pertenencia_etnica)
        ws.cell(row=row_num, column=13, value=usuario.codigo_ocupacion)
        ws.cell(row=row_num, column=14, value=usuario.codigo_pais)  # Placeholder para país si no está implementado
        ws.cell(row=row_num, column=15, value=usuario.codigo_nivel_educativo)

    # Datos de Gestación
        gestacion = Gestacion.objects.filter(consecutivo_registro__id_usuario=usuario).first()
        ws.cell(row=row_num, column=16, value=gestacion.gestante if gestacion else "N/A")
        ws.cell(row=row_num, column=17, value=gestacion.acido_folico_preconcepcional if gestacion else "N/A")
        ws.cell(row=row_num, column=18, value=gestacion.fecha_probable_parto.strftime('%Y-%m-%d') if gestacion and gestacion.fecha_probable_parto else "N/A")
        ws.cell(row=row_num, column=19, value=gestacion.clasificacion_riesgo_gestacional if gestacion else "N/A")
        ws.cell(row=row_num, column=20, value=gestacion.fecha_atencion_parto.strftime('%Y-%m-%d') if gestacion and gestacion.fecha_atencion_parto else "N/A")
        ws.cell(row=row_num, column=21, value=gestacion.fecha_salida_atencion_parto.strftime('%Y-%m-%d') if gestacion and gestacion.fecha_salida_atencion_parto else "N/A")
        ws.cell(row=row_num, column=22, value=gestacion.fecha_primera_consulta_prenatal.strftime('%Y-%m-%d') if gestacion and gestacion.fecha_primera_consulta_prenatal else "N/A")
        ws.cell(row=row_num, column=23, value=gestacion.fecha_ultimo_control_prenatal.strftime('%Y-%m-%d') if gestacion and gestacion.fecha_ultimo_control_prenatal else "N/A")
        ws.cell(row=row_num, column=24, value=gestacion.suministro_acido_folico if gestacion else "N/A")
        ws.cell(row=row_num, column=25, value=gestacion.suministro_sulfato_ferroso if gestacion else "N/A")
        ws.cell(row=row_num, column=26, value=gestacion.suministro_carbonato_calcio if gestacion else "N/A")

    # Datos de Agudeza Visual
        agudeza_visual = AgudezaVisual.objects.filter(consecutivo_registro__id_usuario=usuario).first()
        ws.cell(row=row_num, column=27, value=agudeza_visual.agudeza_visual_lejana_ojo_izquierdo if agudeza_visual else "N/A")
        ws.cell(row=row_num, column=28, value=agudeza_visual.agudeza_visual_lejana_ojo_derecho if agudeza_visual else "N/A")
        ws.cell(row=row_num, column=29, value=agudeza_visual.fecha_valoracion_agudeza_visual.strftime('%Y-%m-%d') if agudeza_visual and agudeza_visual.fecha_valoracion_agudeza_visual else "N/A")

    # Datos de Anticoncepcion

        anticoncepcion = Anticoncepcion.objects.filter(consecutivo_registro__id_usuario=usuario).first()

        ws.cell(row=row_num, column=30, value=anticoncepcion.fecha_atencion_asesoria_anticoncep if anticoncepcion else "N/A")
        ws.cell(row=row_num, column=31, value=anticoncepcion.suministro_metodo_anticoceptivo if anticoncepcion else "N/A")
        ws.cell(row=row_num, column=32, value=anticoncepcion.fecha_sumistro_ant if anticoncepcion else "N/A")

    # Datos de Ca_Cervix

        ca_cervix = Ca_Cervix.objects.filter(consecutivo_registro__id_usuario=usuario).first()

        ws.cell(row=row_num, column=33, value=ca_cervix.tamizaje_cancer_cuello_uterino if ca_cervix else "N/A")
        ws.cell(row=row_num, column=34, value=ca_cervix.tratamiento_ablativo_escision if ca_cervix else "N/A")
        ws.cell(row=row_num, column=35, value=ca_cervix.fecha_tamizaje_cancer_cuello_uterino if ca_cervix else "N/A")
        ws.cell(row=row_num, column=36, value=ca_cervix.resultado_tamizaje_cancer_cuello_uterino if ca_cervix else "N/A")
        ws.cell(row=row_num, column=37, value=ca_cervix.calidad_muestra_citologia if ca_cervix else "N/A")
        ws.cell(row=row_num, column=38, value=ca_cervix.codigo_habilitacion_ips_cancer_cuello_uterino if ca_cervix else "N/A")
        ws.cell(row=row_num, column=39, value=ca_cervix.fecha_colposcopia if ca_cervix else "N/A")
        ws.cell(row=row_num, column=40, value=ca_cervix.fecha_biopsia_cervicouterina if ca_cervix else "N/A")
        ws.cell(row=row_num, column=41, value=ca_cervix.resultado_biopsia_cervicouterina if ca_cervix else "N/A")

    #Datos Gestación menores hasta 7 meses

        gestacion_menores=Gestacion_menores_hasta_siente_meses.objects.filter(consecutivo_registro__id_usuario=usuario).first()

        ws.cell(row=row_num, column=42, value=gestacion_menores.fecha_atencion_promocion_apoyo_lactancia   if  ca_cervix else "N/A")

    #Datos No_Reporte

        no_reportes=No_Reporte.objects.filter(consecutivo_registro__id_usuario=usuario).first()

        ws.cell(row=row_num, column=43, value=no_reportes.sifilis_gestacional_congenita if  no_reportes else "N/A")
        ws.cell(row=row_num, column=44, value=no_reportes.lepra  if  no_reportes else "N/A")
        ws.cell(row=row_num, column=45, value=no_reportes.obesidad_desnutricion  if  no_reportes else "N/A")
        ws.cell(row=row_num, column=46, value=no_reportes.enfermedad_mental  if  no_reportes else "N/A")
        ws.cell(row=row_num, column=47, value=no_reportes.cancer_cervix  if  no_reportes else "N/A")
        ws.cell(row=row_num, column=48, value=no_reportes.dpt_menores  if  no_reportes else "N/A")
        ws.cell(row=row_num, column=49, value=no_reportes.neumococo  if  no_reportes else "N/A")
        ws.cell(row=row_num, column=50, value=no_reportes.consulta_psicologia  if  no_reportes else "N/A")
        ws.cell(row=row_num, column=51, value=no_reportes.preservativos  if  no_reportes else "N/A")
        ws.cell(row=row_num, column=52, value=no_reportes.fecha_hemoglobina  if  no_reportes else "N/A")
        ws.cell(row=row_num, column=53, value=no_reportes.tratamiento_sifilis_gestacional if  no_reportes else "N/A")
        ws.cell(row=row_num, column=54, value=no_reportes.tratamiento_sifilis_congenita  if  no_reportes else "N/A")

    # Datos Primera_Infancia

    primera_infancia= Primera_Infancia.objects.filter(consecutivo_registro__id_usuario=usuario).first()

    ws.cell(row=row_num, column=55, value=primera_infancia.Suministro_fortificacion_casera if  primera_infancia else "N/A")
    ws.cell(row=row_num, column=56, value=primera_infancia.suministro_de_vitamina_A if  primera_infancia else "N/A")
    ws.cell(row=row_num, column=57, value=primera_infancia.suministro_hierro if  primera_infancia else "N/A")

    # Datos Recien_Nacidos

    recien_nacidos= Recien_Nacidos.objects.filter(consecutivo_registro__id_usuario=usuario).first()

    ws.cell(row=row_num, column=58, value=recien_nacidos.resultado_tamizaje_auditivo_neonatal if  recien_nacidos else "N/A")
    ws.cell(row=row_num, column=59, value=recien_nacidos.resultado_tamizaje_visual_neonatal if  recien_nacidos else "N/A")
    ws.cell(row=row_num, column=60, value=recien_nacidos.resultado_oximetria_pre_post_ductal if  recien_nacidos else "N/A")
    ws.cell(row=row_num, column=61, value=recien_nacidos.fecha_oximetria_pre_post_ductal if  recien_nacidos else "N/A")
    ws.cell(row=row_num, column=62, value=recien_nacidos.fecha_tamizaje_auditivo_neonatal if  recien_nacidos else "N/A")
    ws.cell(row=row_num, column=63, value=recien_nacidos.fecha_tamizaje_visual_neonatal if  recien_nacidos else "N/A")
    ws.cell(row=row_num, column=64, value=recien_nacidos.fecha_tsh_neonatal if  recien_nacidos else "N/A")
    ws.cell(row=row_num, column=65, value=recien_nacidos.resultado_tsh_neonatal if  recien_nacidos else "N/A")

   # Datos Riesgo_Cardiovascular

    riesgo_cardiovascular= Riesgo_Cardiovascular.objects.filter(consecutivo_registro__id_usuario=usuario).first()

    ws.cell(row=row_num, column=66, value=riesgo_cardiovascular.consumo_tabaco if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=67, value=riesgo_cardiovascular.resultado_glicemia_basal if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=68, value=riesgo_cardiovascular.fecha_toma_ldl if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=69, value=riesgo_cardiovascular.resultado_ldl if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=70, value=riesgo_cardiovascular.resultado_hdl if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=71, value=riesgo_cardiovascular.resultado_trigliceridos if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=72, value=riesgo_cardiovascular.fecha_toma_hemoglobina if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=73, value=riesgo_cardiovascular.resultado_hemoglobina if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=74, value=riesgo_cardiovascular.fecha_toma_glicemia_basal if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=75, value=riesgo_cardiovascular.fecha_toma_creatinina if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=76, value=riesgo_cardiovascular.resultado_creatinina if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=77, value=riesgo_cardiovascular.fecha_toma_hdl if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=78, value=riesgo_cardiovascular.clasificacion_riesgo_cardiovascular if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=79, value=riesgo_cardiovascular.clasificacion_riesgo_metabolico if  riesgo_cardiovascular else "N/A")
    ws.cell(row=row_num, column=80, value=riesgo_cardiovascular.fecha_toma_trigliceridos if  riesgo_cardiovascular else "N/A")

    # Datos Test_0_12

    test_0_12= Test_0_12.objects.filter(consecutivo_registro__id_usuario=usuario).first()

    ws.cell(row=row_num, column=81, value=test_0_12.resultado_tamizaje_vale if  test_0_12 else "N/A")
    ws.cell(row=row_num, column=82, value=test_0_12.fecha_tamizaje_vale if  test_0_12 else "N/A")

    # Datos Test_0_7

    test_0_7= Test_0_7.objects.filter(consecutivo_registro__id_usuario=usuario).first()

    ws.cell(row=row_num, column=83, value=test_0_7.resultado_motricidad_gruesa if  test_0_7 else "N/A")
    ws.cell(row=row_num, column=84, value=test_0_7.resultado_motricidad_finoadaptativa if  test_0_7 else "N/A")
    ws.cell(row=row_num, column=85, value=test_0_7.resultado_personal_social if  test_0_7 else "N/A")
    ws.cell(row=row_num, column=86, value=test_0_7.resultado_audicion_lenguaje if  test_0_7 else "N/A")

    # Datos Test_Mayores_60

    test_mayores_60 = Test_Mayores_60.objects.filter(consecutivo_registro__id_usuario=usuario).first()

    ws.cell(row=row_num, column=87, value=test_mayores_60.test_vejez  if  test_mayores_60  else "N/A")

    # Datos Toda_poblacion

    toda_poblacion = Toda_poblacion.objects.filter(consecutivo_registro__id_usuario=usuario).first()

    ws.cell(row=row_num, column=88, value= toda_poblacion.fecha_peso  if   toda_poblacion else "N/A")
    ws.cell(row=row_num, column=89, value= toda_poblacion.peso  if   toda_poblacion else "N/A")
    ws.cell(row=row_num, column=90, value= toda_poblacion.fecha_talla  if   toda_poblacion else "N/A")
    ws.cell(row=row_num, column=91, value= toda_poblacion.talla  if   toda_poblacion else "N/A")
    ws.cell(row=row_num, column=92, value= toda_poblacion.resultado_tamizaje_hepatitis_c  if   toda_poblacion else "N/A")
    ws.cell(row=row_num, column=93, value= toda_poblacion.fecha_valoracion_integral  if   toda_poblacion else "N/A")
    ws.cell(row=row_num, column=94, value= toda_poblacion.fecha_antigeno_hepatitis_b  if   toda_poblacion else "N/A")
    ws.cell(row=row_num, column=95, value= toda_poblacion.resultado_antigeno_hepatitis_b  if   toda_poblacion else "N/A")
    ws.cell(row=row_num, column=96, value= toda_poblacion.fecha_tamizaje_sifilis  if   toda_poblacion else "N/A")
    ws.cell(row=row_num, column=97, value= toda_poblacion.resultado_tamizaje_sifilis  if   toda_poblacion else "N/A")
    ws.cell(row=row_num, column=98, value= toda_poblacion.fecha_prueba_vih  if   toda_poblacion else "N/A")
    ws.cell(row=row_num, column=99, value= toda_poblacion.resultado_prueba_vih  if   toda_poblacion else "N/A")
    ws.cell(row=row_num, column=100, value= toda_poblacion.fecha_tamizaje_hepatitis_c  if   toda_poblacion else "N/A")

# Datos Tuberculosis

    tuberculosis = Tuberculosis.objects.filter(consecutivo_registro__id_usuario=usuario).first()

    ws.cell(row=row_num, column=101, value= tuberculosis.Sintomatico_respiratorio  if   tuberculosis else "N/A")
    ws.cell(row=row_num, column=102, value= tuberculosis.Fecha_baciloscopia  if   tuberculosis else "N/A")
    ws.cell(row=row_num, column=103, value= tuberculosis.Resultado_baciloscopia  if   tuberculosis else "N/A")

#Datos Cancer de colon

    colon=Ca_Colon.objects.filter(consecutivo_registro__id_usuario=usuario).first()

    ws.cell(row=row_num, column=104, value= colon.resultado_prueba_sangre_oculta_mf if   colon else "N/A")
    ws.cell(row=row_num, column=105, value= colon.resultado_colonoscopia_tamizaje  if   colon else "N/A")
    ws.cell(row=row_num, column=106, value= colon.fecha_colonoscopia_tamizaje  if   colon else "N/A")
    ws.cell(row=row_num, column=107, value= colon.fecha_prueba_sangre_oculta  if   colon else "N/A")

#Campos  Cancer Prostata

    ca_prostata=Ca_Prostata.objects.filter(consecutivo_registro__id_usuario=usuario).first()

    ws.cell(row=row_num, column=108, value= ca_prostata.resultado_tacto_rectal if ca_prostata else "N/A")
    ws.cell(row=row_num, column=109, value= ca_prostata.fecha_tacto_rectal if ca_prostata else "N/A")
    ws.cell(row=row_num, column=110, value= ca_prostata.fecha_toma_psa if ca_prostata else "N/A")
    ws.cell(row=row_num, column=111, value= ca_prostata.resultado_psa if ca_prostata else "N/A")

#datos Salud Bucal

    salud_bucal=Salud_Bucal.objects.filter(consecutivo_registro__id_usuario=usuario).first()

    ws.cell(row=row_num, column=112, value=salud_bucal.fecha_atencion_salud_bucal if salud_bucal else "N/A")
    ws.cell(row=row_num, column=113, value=salud_bucal.cop_por_persona if salud_bucal else "N/A")

#Datos cancer de mama

    ca_mama=Ca_Mama.objects.filter(consecutivo_registro__id_usuario=usuario).first()

    ws.cell(row=row_num, column=114, value=ca_mama.fecha_toma_mamografia if ca_mama else "N/A")
    ws.cell(row=row_num, column=115, value=ca_mama.resultado_mamografia if ca_mama else "N/A")
    ws.cell(row=row_num, column=116, value=ca_mama.fecha_toma_biopsia_mama if ca_mama else "N/A")
    ws.cell(row=row_num, column=117, value=ca_mama.fecha_resultado_biopsia_mama if ca_mama else "N/A")
    ws.cell(row=row_num, column=118, value=ca_mama.resultado_biopsia_mama if ca_mama else "N/A")

    # Ajustar las columnas automáticamente
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Configurar la respuesta HTTP para descargar el archivo
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_usuarios_completo.xlsx"'
    wb.save(response)
    return response